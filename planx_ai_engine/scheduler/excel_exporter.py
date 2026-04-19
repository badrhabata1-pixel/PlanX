from __future__ import annotations

from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_schedule_to_excel(
    rows: List[Dict[str, Any]],
    evaluation: Any,
    output_path: str = "planx_final_schedule.xlsx",
) -> str:
    wb = Workbook()

    # -----------------------------
    # Sheet 1: Final Schedule
    # -----------------------------
    ws = wb.active
    ws.title = "Final Schedule"
    ws.sheet_view.rightToLeft = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "PlanX Final Exam Schedule"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = "Date | Time | Course Title | Students Numbers | Exam Type | Places"
    ws["A2"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Date",
        "Day",
        "Time",
        "Course Code",
        "Course Title",
        "Students Numbers",
        "Exam Type",
        "Places",
    ]

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    _style_table_header(ws, header_row, len(headers))

    current_row = header_row + 1
    previous_date = None

    for row in rows:
        # date separator style
        if previous_date is not None and row["date"] != previous_date:
            current_row += 1

        ws.cell(row=current_row, column=1, value=row["date"])
        ws.cell(row=current_row, column=2, value=row["day_name"])
        ws.cell(row=current_row, column=3, value=row["time"])
        ws.cell(row=current_row, column=4, value=row["course_code"])
        ws.cell(row=current_row, column=5, value=row["course_title"])
        ws.cell(row=current_row, column=6, value=row["students_numbers"])
        ws.cell(row=current_row, column=7, value=row["exam_type"])
        ws.cell(row=current_row, column=8, value=row["places"])

        _style_schedule_row(ws, current_row, row["date"] == previous_date)

        previous_date = row["date"]
        current_row += 1

    # Freeze header
    ws.freeze_panes = "A5"

    # Column widths
    column_widths = {
        "A": 14,
        "B": 14,
        "C": 18,
        "D": 16,
        "E": 42,
        "F": 18,
        "G": 16,
        "H": 55,
    }
    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    # Alignments
    for row in ws.iter_rows(min_row=5, max_row=current_row - 1, min_col=1, max_col=8):
        for cell in row:
            if cell.column in [5, 8]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -----------------------------
    # Sheet 2: Summary
    # -----------------------------
    summary_ws = wb.create_sheet(title="Summary")

    summary_ws["A1"] = "PlanX Final Summary"
    summary_ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    summary_ws["A1"].fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.merge_cells("A1:B1")

    summary_rows = [
        ("Assigned Exams", evaluation.summary.get("assigned_exams")),
        ("Unassigned Exams", evaluation.summary.get("unassigned_exams")),
        ("Component Count", evaluation.summary.get("component_count")),
        ("Feasible", evaluation.feasible),
        ("Final Total Penalty", evaluation.total_penalty),
    ]

    start_row = 3
    for index, (label, value) in enumerate(summary_rows, start=start_row):
        summary_ws[f"A{index}"] = label
        summary_ws[f"B{index}"] = value

    summary_ws["A10"] = "Penalty Breakdown"
    summary_ws["A10"].font = Font(bold=True, size=12, color="FFFFFF")
    summary_ws["A10"].fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")

    current_row = 11
    for name, penalty in evaluation.component_penalties.items():
        summary_ws[f"A{current_row}"] = name
        summary_ws[f"B{current_row}"] = penalty
        current_row += 1

    _style_summary_sheet(summary_ws, current_row)
    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 18

    wb.save(output_path)
    return output_path


def _style_table_header(ws, header_row: int, header_count: int) -> None:
    fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    font = Font(bold=True, color="000000")
    alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, header_count + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _style_schedule_row(ws, row_idx: int, same_as_previous_date: bool) -> None:
    thin = Side(style="thin", color="B7B7B7")
    medium = Side(style="medium", color="7F7F7F")

    fill = PatternFill(
        fill_type="solid",
        start_color="FFFFFF" if same_as_previous_date else "F7FBFF",
        end_color="FFFFFF" if same_as_previous_date else "F7FBFF",
    )

    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.border = Border(
            left=thin,
            right=thin,
            top=medium if not same_as_previous_date else thin,
            bottom=thin,
        )


def _style_summary_sheet(ws, last_row: int) -> None:
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=3, max_row=last_row - 1, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")